# import xlrd
#
# rb = xlrd.open_workbook('d:/schedule.xlsx')
# sheet = rb.sheet_by_index(0)
# for rownum in range(sheet.nrows):
#     row = sheet.row_values(rownum)
#     for c_el in row:
#         print(c_el)
# import pandas as pd
#
# schedule1 = pd.read_excel('d:/schedule.xlsx', index_col='День недели', sheet_name='10',
#                           usecols=['День недели', '№Урока', '10мед', 'Кабинет'])
#
# print(schedule1)
import logging

from openpyxl import load_workbook

logger = logging.getLogger()
wb = load_workbook('d:/schedule.xlsx')
sheet = wb['11П']

rows = sheet.max_row
cols = sheet.max_column
timetable = []
for i in range(1, rows + 1):
    string = ''
    for j in range(1, 5):
        cell = sheet.cell(row=i, column=j)
        string = string + str(cell.value) + ','
    timetable.append(string)
    # print(string)
# timetable.pop(0)
print(timetable)  # получаем список, где каждый элемент это 1 урок
# logger.error('TEST')
for item in timetable:
    data = item.split(',')
    # print(data)
    data1 = f'{data[0]} {data[1]} {data[2]} {data[3]}'  # получаем расписание вида: "1 математика 22"
    print(data1)

    # а дальше я не знаю...
# for row in data1:
#     pass
