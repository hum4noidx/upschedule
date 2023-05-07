import logging
from enum import IntEnum

from openpyxl.reader.excel import load_workbook

from tgbot.models.lesson import Lesson

logger = logging.getLogger(__name__)


class _ColumnDataType(IntEnum):
    """
        Used to determine the type of data in the column. The order of the values is important.

        The values are based on the column offset from the group name column:
        +-------------+--------+---------+
        |             |        | Класс  |
        +-------------+--------+---------+
        | День недели | № Урока | Предмет |
        | -2          | -1     | 0       |
        +-------------+--------+---------+
        """
    SUBJECT = 0
    LESSON_NUMBER = -1
    WEEKDAY = -2


class Parser:
    def __init__(self, filename: str):
        self._filename = filename

    def load_workbook(self):
        return load_workbook(self._filename)

    def find_sheets(self):
        workbook = self.load_workbook()
        sheets = []
        for sheet in workbook.worksheets:
            sheets.append(sheet)
        return sheets

    def find_groups(self):
        workbook = self.load_workbook()
        groups = []
        for sheet in workbook.worksheets:
            for col in sheet.iter_rows(
                    min_col=1,
                    min_row=2,
                    max_row=5,
            ):
                print(col)
                for cell in col:
                    print(cell.value)
                    if cell.value is not None:
                        groups.append(cell.value)
        return set(groups)

    def filter_groups(self):
        groups = self.find_groups()
        filtered_groups = []
        for group in groups:
            if any(char.isdigit() for char in str(group)):
                if not str(group).isdigit():
                    filtered_groups.append(group)
        return filtered_groups

    def get_group_name_column(self, group_name: str, worksheet: str):
        workbook = self.load_workbook()
        sheet = workbook[worksheet]
        for col in sheet.iter_cols(min_col=1, max_col=15, min_row=2, max_row=5):
            for cell in col:
                if cell.value == group_name:
                    return cell
        raise ValueError(f'Group name column not found for group {group_name}')

    def get_lesson_number_rows(self, worksheet: str):
        workbook = self.load_workbook()
        sheet = workbook[worksheet]
        lesson_number_rows = []
        for row in sheet.iter_rows(min_col=3, max_col=3, min_row=6, max_row=50):
            for cell in row:
                if cell.value is not None:
                    lesson_number_rows.append({cell.value: cell.row})
        return lesson_number_rows

    def get_lessons(self, group_name: str, worksheet: str):
        workbook = self.load_workbook()
        sheet = workbook[worksheet]
        class_name_column = self.get_group_name_column(group_name, worksheet)
        lessons = []
        last_lesson = 0
        day_number = 1
        for row in sheet.iter_rows(
                min_col=class_name_column.column - 1, max_col=class_name_column.column + 1,
                min_row=class_name_column.row + 3, max_row=class_name_column.row + 1 + 50,
                values_only=True
        ):
            lesson_number = row[0]
            subject = row[1]
            room = row[2]
            if lesson_number is None:
                continue
            if subject is None:
                subject = 'Окно'
            try:
                if int(lesson_number) < last_lesson:
                    day_number += 1
            except ValueError:
                if lesson_number == '9-10' or '8-10':
                    day_number += 1
            lessons.append(
                Lesson(
                    day_of_week=day_number,
                    lesson_number=lesson_number,
                    subject=subject,
                    room=room
                )
            )
        return lessons
